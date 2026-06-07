"""
Yoga Center Attendance & Fee Status Dashboard
Production-Ready Version with Enhanced Error Handling & Bug Fixes
"""
import os
import re
import tempfile
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Tuple, List, Dict
from datetime import datetime

import streamlit as st
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION & CONSTANTS
# ============================================================================

PAGE_SIZE = A4
LEFT_MARGIN, RIGHT_MARGIN = 14 * mm, 14 * mm
TOP_MARGIN, BOTTOM_MARGIN = 20 * mm, 20 * mm

# Brand Colors
BRAND_RED = colors.HexColor("#8A1E1E")
BRAND_DARK = colors.HexColor("#222222")
BRAND_GREY = colors.HexColor("#666666")
BRAND_LIGHT = colors.HexColor("#F9F9F9")
BRAND_GREEN = colors.HexColor("#2E7D32")
CHART_RED = "#8A1E1E"
CHART_GREEN = "#2E7D32"

# Day of Week Order
DOW_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# ============================================================================
# CUSTOM EXCEPTIONS
# ============================================================================

class ReportValidationError(Exception):
    """Custom exception for report validation errors with user-friendly messages"""
    pass

# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class ReportMeta:
    center_name: str
    report_month: str
    report_period: str
    month_key: str
    date_start: datetime
    year_str: str

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def safe_str(x) -> str:
    """Convert any value to safe string, handling NaN/None"""
    if pd.isna(x) or x is None:
        return ""
    return str(x).strip()

def normalize_col(col) -> str:
    """Normalize column names to consistent format"""
    if isinstance(col, pd.Timestamp):
        return col.strftime("%d-%b-%Y")
    if hasattr(col, "year") and hasattr(col, "month") and hasattr(col, "day"):
        try:
            return pd.Timestamp(col).strftime("%d-%b-%Y")
        except:
            return str(col).strip()
    return str(col).strip()

def clean_center_name(raw_center: str) -> str:
    """Standardize center names for consistent reporting"""
    txt = safe_str(raw_center)
    if not txt:
        return "UNKNOWN CENTER"
    
    if "-" in txt:
        txt = txt.split("-", 1)[-1].strip()
    
    txt = re.sub(r"\s*\(.*?\)\s*", "", txt).strip()
    
    lower = txt.lower().replace(" ", "")
    replacements = {
        "vijayanagar(blore)": "VIJAYANAGARA",
        "vijayanagar": "VIJAYANAGARA",
        "sadashivanagar": "SADASHIVANAGARA",
        "gayathrinagar": "GAYATHRI NAGAR",
        "gyn": "GAYATHRI NAGAR",
        "rysri": "RYSRI"
    }
    
    for k, v in replacements.items():
        if lower == k.replace(" ", ""):
            return v
    
    return txt.upper()

def clean_batch_name(raw_name: str) -> str:
    """Standardize batch names for reporting"""
    txt = safe_str(raw_name)
    # FIX: Handle both comma and space separated duplicates (e.g., "General,General")
    txt = re.sub(r'(?i)(general)[,\s]+general', r'\1', txt)
    txt = re.sub(r'(?i)(therapy)[,\s]+therapy', r'\1', txt)
    txt = re.sub(r'(?i)(junior)[,\s]+junior', r'\1', txt)
    txt = re.sub(r'(?i)(senior)[,\s]+senior', r'\1', txt)
    txt = re.sub(r'(?i)children\s*yoga.*children\s*yoga.*', 'Children Yoga', txt)
    
    # Remove suffixes like YG, YC
    txt = re.sub(r'\s+(YG|YC|WE)\s*$', '', txt, flags=re.IGNORECASE)
    return txt.strip()

def extract_numeric_id(val) -> str:
    """Extract numeric student ID from various formats"""
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
    # Find trailing numbers (handles "9-GYN/Y/4162" → "4162")
    match = re.search(r'(\d+)$', s)
    v = match.group(1) if match else s
    return v.lstrip('0') or '0'

def parse_mixed_date(val):
    """
    Parse dates from multiple formats found in fee reports.
    FIX: Prioritizes DD/MM/YY over MM/DD/YY to prevent Indian dates from being read as US dates.
    """
    if pd.isna(val) or val is None or safe_str(val) == "":
        return None
    
    # If it's already a datetime object from Excel, preserve it
    if isinstance(val, (pd.Timestamp, datetime)):
        return val
        
    val_str = str(val).strip()
    
    # Remove time component if present
    if " " in val_str:
        val_str = val_str.split(" ")[0]
    
    # FIX: Indian/UK formats first, then US formats as fallback
    formats = [
        "%d/%m/%Y",      # 13/05/2026
        "%d-%m-%Y",      # 13-05-2026
        "%Y-%m-%d",      # 2026-05-13
        "%d/%m/%y",      # 1/4/26 (India format - PRIORITIZED)
        "%d-%m-%y",      # 1-4-26
        "%m/%d/%y",      # 1/4/26 (US format - FALLBACK)
        "%m-%d-%y",      # 1-4-26
    ]
    
    for fmt in formats:
        try:
            return pd.to_datetime(val_str, format=fmt)
        except:
            continue
    
    # Final fallback - let pandas infer
    try:
        return pd.to_datetime(val, errors='coerce')
    except:
        return None

def find_header_row(path: str, target_cols: list, max_rows: int = 50) -> int:
    """Find the header row by scanning for target column names."""
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        for r in range(1, min(max_rows, ws.max_row) + 1):
            values = [safe_str(ws.cell(r, c).value).lower() for c in range(1, min(ws.max_column, 30) + 1)]
            matches = sum(1 for t in target_cols if any(t.lower() in v for v in values))
            if matches >= 2:
                return r - 1  # Return 0-indexed
        
        return -1
    except Exception:
        return -1

def attendance_cell_to_bool(series: pd.Series) -> pd.Series:
    """Convert attendance cell values to boolean (Present=True)"""
    s = series.astype(str).str.strip().str.lower()
    present_values = {"present", "p", "x", "1", "yes", "y", "true"}
    return s.isin(present_values)

def extract_metadata_from_attendance(raw: pd.DataFrame) -> str:
    """Extract center name from attendance file metadata"""
    for _, row in raw.head(10).iterrows():
        vals = [safe_str(v) for v in row.tolist()[:6]]
        joined = " | ".join(v for v in vals if v)
        if any(k in joined.lower() for k in ["vijayanagar", "sadashivanagar", "jayanagar", "rysri", "gayathri"]):
            for v in vals:
                if v and not re.search(r"^\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}$", v):
                    if any(k in v.lower() for k in ["vijayanagar", "sadashivanagar", "jayanagar", "rysri", "gayathri"]):
                        return v
    return ""

def format_mobile_number(mobile: str) -> str:
    """Standardize mobile number format"""
    if not mobile or mobile in ["N/A", "NaN", "nan", ""]:
        return "Not provided"
    
    cleaned = re.sub(r'[\s\-\(\)]', '', str(mobile))
    if len(cleaned) > 10:
        cleaned = cleaned[-10:]
    
    if re.match(r'^\d{10}$', cleaned):
        return cleaned
    
    return str(mobile).strip()

def extract_coverage_months(particulars: str) -> str:
    """Extract month tags from Particulars field"""
    if not particulars:
        return "None"
    
    months = re.findall(r'\[[A-Za-z]{3}[\s\-/]*\d{2}\]', particulars, re.IGNORECASE)
    
    if months:
        return ", ".join(months)
    
    return "None"

# ============================================================================
# DATA LOADING FUNCTIONS
# ============================================================================

def read_fee_report(input_file: str) -> pd.DataFrame:
    """Load and clean fee report data."""
    header_row = find_header_row(input_file, ["stud id", "stud name", "mobile no", "particulars"], max_rows=50)
    
    if header_row == -1:
        raise ReportValidationError(
            "⚠️ Could not locate the header row in the Fee Report. "
            "Please ensure the file contains columns like 'Stud ID', 'Stud Name', 'Mobile No', and 'Particulars'."
        )
    
    raw = pd.read_excel(input_file, header=header_row)
    raw.columns = [str(c).strip() for c in raw.columns]
    
    id_col = next((c for c in raw.columns if "stud id" in c.lower() or "student id" in c.lower()), None)
    status_col = next((c for c in raw.columns if "status" in c.lower()), None)
    mobile_col = next((c for c in raw.columns if "mobile" in c.lower()), None)
    particulars_col = next((c for c in raw.columns if "particulars" in c.lower()), None)
    pymt_dt_col = next((c for c in raw.columns if "pymt dt" in c.lower() or "payment date" in c.lower()), None)
    
    if not id_col:
        raise ReportValidationError("⚠️ Fee report missing 'Stud ID' column.")
    if not status_col:
        raise ReportValidationError("⚠️ Fee report missing 'Status' column.")
    
    fee_records = []
    for idx, row in raw.iterrows():
        val_id = row.get(id_col)
        if pd.notna(val_id):
            num_id = extract_numeric_id(val_id)
            status = safe_str(row.get(status_col, "")).lower()
            
            mobile = "Not provided"
            if mobile_col and pd.notna(row.get(mobile_col)):
                m = row[mobile_col]
                if isinstance(m, (int, float)):
                    try: mobile = str(int(m))
                    except: mobile = "Not provided"
                else:
                    mobile = format_mobile_number(str(m))
            
            particulars = safe_str(row.get(particulars_col, "")) if particulars_col else ""
            
            pymt_dt = None
            if pymt_dt_col and pd.notna(row.get(pymt_dt_col)):
                pymt_dt = parse_mixed_date(row[pymt_dt_col])
            
            fee_records.append({
                "Norm_ID": num_id,
                "Status": status,
                "Mobile": mobile,
                "Particulars": particulars,
                "Pymt_Dt": pymt_dt
            })
    
    if len(fee_records) == 0:
        raise ReportValidationError("⚠️ No valid fee records found.")
    
    return pd.DataFrame(fee_records)

def read_and_clean_attendance(input_file: str):
    """Load and clean attendance data with smart date detection."""
    try:
        wb = load_workbook(input_file, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        sample_text = ""
        for r in range(1, min(15, ws.max_row) + 1):
            sample_text += " ".join([safe_str(ws.cell(r, c).value).lower() for c in range(1, min(ws.max_column, 25) + 1)])
        
        if "admission fee" in sample_text or "subscription fees" in sample_text or "particulars" in sample_text:
            raise ReportValidationError(
                "⚠️ It looks like you uploaded the Fee Report in the Attendance slot. Please swap the files."
            )
    except Exception:
        pass
    
    header_row = find_header_row(input_file, ["studentname", "studentid", "slno", "batch"], max_rows=50)
    
    if header_row == -1:
        raise ReportValidationError(
            "⚠️ Invalid Attendance File format. Could not find header columns like 'StudentId', 'StudentName', or 'Batch'."
        )
    
    raw = pd.read_excel(input_file, header=header_row)
    raw.columns = [normalize_col(c) for c in raw.columns]
    raw = raw.dropna(how="all")
    
    if "StudentId" in raw.columns:
        raw = raw.dropna(subset=["StudentId"])
        raw['data_points'] = raw.notna().sum(axis=1)
        raw = raw.sort_values('data_points', ascending=False).drop_duplicates(subset=['StudentId'], keep='first')
        raw = raw.drop(columns=['data_points'])
    
    raw = raw.reset_index(drop=True)
    
    if len(raw) == 0:
        raise ReportValidationError("⚠️ No valid student records found after cleaning the attendance data.")
    
    status_idx = raw.columns.get_loc("Status") if "Status" in raw.columns else -1
    start_col = status_idx + 1 if status_idx != -1 else 0
    
    # FIX: Smart date format detection for attendance columns
    sample_cols = [c for c in raw.columns[start_col:start_col+15] if safe_str(c).strip().lower() not in {"present", "absent"}]
    is_dmy = False
    for c in sample_cols:
        parts = safe_str(c).split("/")
        if len(parts) >= 2:
            try:
                p1, p2 = int(parts[0]), int(parts[1])
                if p1 > 12:
                    is_dmy = True
                    break
                elif p2 > 12:
                    is_dmy = False
                    break
            except:
                pass
    
    date_fmts = ["%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d/%b/%y"] if is_dmy else \
                ["%m/%d/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d/%b/%y"]

    date_cols, date_index = [], []
    for c in raw.columns[start_col:]:
        col_str = safe_str(c).strip().lower()
        if col_str in {"present", "absent"}:
            continue
        
        dt = None
        for fmt in date_fmts:
            try:
                dt = pd.to_datetime(c, format=fmt)
                break
            except:
                continue
        
        if pd.notna(dt):
            date_cols.append(c)
            date_index.append(dt)
        else:
            dt2 = pd.to_datetime(c, errors="coerce")
            if pd.notna(dt2):
                date_cols.append(c)
                date_index.append(dt2)
    
    if not date_cols:
        raise ReportValidationError("⚠️ No valid attendance dates were found in this file.")
    
    sort_idx = np.argsort(pd.to_datetime(date_index))
    date_cols = [date_cols[i] for i in sort_idx]
    date_index = pd.DatetimeIndex([date_index[i] for i in sort_idx])
    
    df = raw.copy()
    for col in ["Center", "Status", "StudentName", "StudentId"]:
        if col in df.columns:
            df[col] = df[col].map(safe_str)
    
    if "Batch" in df.columns:
        df["Batch"] = df["Batch"].apply(clean_batch_name)
    
    presence = pd.DataFrame({c: attendance_cell_to_bool(df[c]) for c in date_cols}, index=df.index)
    df["Present_Count_Calc"] = presence.sum(axis=1).astype(int)
    df["System_ID"] = df["StudentId"].apply(extract_numeric_id)
    
    center_raw = ""
    if "Center" in df.columns and df["Center"].str.strip().ne("").any():
        center_raw = df["Center"].mode().iloc[0] if len(df["Center"].mode()) > 0 else ""
    
    if not center_raw:
        center_raw = extract_metadata_from_attendance(raw)
    
    date_start, date_end = date_index.min(), date_index.max()
    
    meta = ReportMeta(
        center_name=clean_center_name(center_raw),
        report_month=date_start.strftime("%B %Y"),
        report_period=f"{date_start.strftime('%d-%b-%Y')} to {date_end.strftime('%d-%b-%Y')}",
        month_key=date_start.strftime("%b_%Y"),
        date_start=date_start,
        year_str=date_start.strftime('%y')
    )
    
    return df, date_cols, date_index, meta, presence

# ============================================================================
# RECONCILIATION LOGIC
# ============================================================================

def reconcile_fees(df: pd.DataFrame, fee_df: pd.DataFrame, target_month: str, year_str: str):
    """Reconcile attendance with fee data to identify unpaid attendees."""
    if fee_df is None or fee_df.empty:
        return set(), df, pd.DataFrame()
    
    active_fees = fee_df[fee_df['Status'] == 'active'].copy()
    if active_fees.empty:
        return set(), df, pd.DataFrame()
    
    month_variants = [
        rf"\[{target_month}[\s\-/]*{year_str}\]",
        rf"\[{target_month[:3]}[\s\-/]*{year_str}\]"
    ]
    combined_pattern = "|".join(month_variants)
    
    active_fees['Covers_Target_Month'] = active_fees['Particulars'].str.contains(
        combined_pattern, case=False, na=False, regex=True
    )
    
    valid_fees = active_fees[active_fees['Covers_Target_Month']].copy()
    
    if valid_fees.empty:
        valid_fees = pd.DataFrame(columns=active_fees.columns)
    
    if not valid_fees.empty:
        valid_fees = valid_fees.sort_values(by='Pymt_Dt', ascending=False, na_position='last')
        latest_valid_fees = valid_fees.drop_duplicates(subset=['Norm_ID'], keep='first')
        paid_ids = set(latest_valid_fees['Norm_ID'].tolist())
    else:
        paid_ids = set()
    
    all_active_fees_sorted = active_fees.sort_values(by='Pymt_Dt', ascending=False, na_position='last')
    latest_any_fees = all_active_fees_sorted.drop_duplicates(subset=['Norm_ID'], keep='first')
    
    mobile_map = latest_any_fees.set_index('Norm_ID')['Mobile'].to_dict()
    pymt_dt_map = latest_any_fees.set_index('Norm_ID')['Pymt_Dt'].to_dict()
    coverage_map = latest_any_fees.set_index('Norm_ID')['Particulars'].to_dict()
    
    df['Mobile No'] = df['System_ID'].map(mobile_map).fillna("Not provided")
    
    def format_date(d):
        if pd.isna(d) or d is None: return "Never"
        try: return d.strftime("%d-%b-%Y")
        except: return "Never"
    
    df['Last Payment Date'] = df['System_ID'].map(pymt_dt_map).apply(format_date)
    df['Subscription Coverage'] = df['System_ID'].map(coverage_map).fillna("").apply(extract_coverage_months)
    
    # Identify unpaid attendees
    attended_mask = df['Present_Count_Calc'] > 0
    not_paid_mask = ~df['System_ID'].isin(paid_ids)
    
    # FIX: Exclude students explicitly marked as 'Inactive' in the attendance register
    active_status_mask = True
    if 'Status' in df.columns:
        active_status_mask = df['Status'].astype(str).str.lower() != 'inactive'
        
    pending_students = df[attended_mask & not_paid_mask & active_status_mask].copy()
    
    cols_to_keep = ['StudentName', 'System_ID', 'Present_Count_Calc', 'Mobile No', 'Last Payment Date', 'Subscription Coverage']
    if 'Status' in pending_students.columns:
        cols_to_keep.insert(2, 'Status')
    
    pending_students = pending_students[cols_to_keep].sort_values(
        by='Present_Count_Calc', ascending=False
    ).reset_index(drop=True)
    
    pending_students.rename(columns={'Present_Count_Calc': 'Days Attended'}, inplace=True)
    
    return paid_ids, df, pending_students

# ============================================================================
# ANALYTICS FUNCTIONS
# ============================================================================

def build_analytics(df, date_cols, date_index, presence, fee_df=None, meta=None):
    """Generate comprehensive analytics from attendance and fee data"""
    
    daily_totals = pd.DataFrame({
        "Date": date_index,
        "Attendance": presence.sum(axis=0).values.astype(int)
    })
    daily_totals["DayOfWeek"] = daily_totals["Date"].dt.day_name()
    daily_totals["ShortDate"] = daily_totals["Date"].dt.strftime("%d-%b")
    
    dow_totals = daily_totals.groupby("DayOfWeek", as_index=False)["Attendance"].sum()
    dow_totals = dow_totals.set_index("DayOfWeek").reindex(DOW_ORDER).fillna(0).reset_index()
    
    if "Batch" in df.columns:
        batch_analysis = df.groupby("Batch", as_index=False).agg(
            Members=("StudentId", "count"),
            Total_Present=("Present_Count_Calc", "sum")
        ).sort_values("Total_Present", ascending=False)
    else:
        batch_analysis = pd.DataFrame(columns=["Batch", "Members", "Total_Present"])
    
    # FIX: Extended upper bound to np.inf to capture all high-frequency attendees safely
    df["Segment"] = pd.cut(
        df["Present_Count_Calc"],
        bins=[-1, 0, 5, 11, np.inf],
        labels=["Inactive (0 visits)", "Occasional (1-5 visits)", "Regular (6-11 visits)", "Dedicated (12+ visits)"]
    )
    segment_counts = df["Segment"].value_counts().reindex(
        ["Inactive (0 visits)", "Occasional (1-5 visits)", "Regular (6-11 visits)", "Dedicated (12+ visits)"]
    ).fillna(0).reset_index()
    segment_counts.columns = ["Member Segment", "Count"]
    
    top_attendees = df[["StudentName", "Batch", "Present_Count_Calc"]].sort_values(
        ["Present_Count_Calc", "StudentName"], ascending=[False, True]
    ).reset_index(drop=True)
    
    least_active = df[["StudentName", "Batch", "Present_Count_Calc"]].sort_values(
        ["Present_Count_Calc", "StudentName"], ascending=[True, True]
    ).reset_index(drop=True)
    
    daily_totals["WeekOfMonth"] = ((daily_totals["Date"].dt.day - 1) // 7) + 1
    weekly_totals = daily_totals.groupby("WeekOfMonth", as_index=False)["Attendance"].sum()
    weekly_totals["Week"] = "Week " + weekly_totals["WeekOfMonth"].astype(str)
    weekly_totals = weekly_totals[["Week", "Attendance"]]
    
    total_att = int(daily_totals["Attendance"].sum())
    avg_visits = round(total_att / max(1, len(df)), 1)
    
    busiest_day = dow_totals.loc[dow_totals['Attendance'].idxmax(), 'DayOfWeek'] if not dow_totals.empty else "N/A"
    
    kpis = pd.DataFrame({
        "Metric": [
            "Registered Members", "Operating Days", "Total Center Visits",
            "Avg Visits per Member", "Busiest Day"
        ],
        "Value": [len(df), len(date_cols), total_att, avg_visits, busiest_day]
    })
    
    insights = [
        f"Habit Building: On average, each member visited {avg_visits} times this month.",
        f"Batch Popularity: The '{batch_analysis.iloc[0]['Batch'] if not batch_analysis.empty else 'N/A'}' batch had the highest footfall." if not batch_analysis.empty else "Batch data not available.",
        f"Peak Activity: {busiest_day} saw the most activity across the month."
    ]
    
    peak_day_row = daily_totals.loc[daily_totals['Attendance'].idxmax()] if not daily_totals.empty else None
    busiest_week = weekly_totals.loc[weekly_totals['Attendance'].idxmax()] if not weekly_totals.empty else None
    
    op_insights = []
    if peak_day_row is not None:
        op_insights.append(f"Highest Single Day: The center saw its peak daily footfall on {peak_day_row['ShortDate']} with {int(peak_day_row['Attendance'])} recorded visits.")
    if busiest_week is not None:
        op_insights.append(f"Busiest Period: {busiest_week['Week']} was the most active operational period, capturing {int(busiest_week['Attendance'])} total visits.")
    
    pending_students = pd.DataFrame()
    fee_summary_text = ""
    
    if fee_df is not None and not fee_df.empty and meta is not None:
        month_str = MONTH_ABBR[meta.date_start.month - 1]
        year_str = meta.year_str
        
        paid_ids, df, pending_students = reconcile_fees(df, fee_df, month_str, year_str)
        
        total_attending = len(df[df['Present_Count_Calc'] > 0])
        total_unpaid = len(pending_students)
        
        if total_attending > 0:
            unpaid_ratio = total_unpaid / total_attending
            if unpaid_ratio > 0.70:
                raise ReportValidationError(
                    f"⚠️ Data Anomaly Detected: {int(unpaid_ratio*100)}% of attending students are flagged as unpaid. "
                    f"Please verify correct files and ID matching."
                )
        
        total_pending_visits = int(pending_students['Days Attended'].sum()) if not pending_students.empty else 0
        total_pending_individuals = len(pending_students)
        
        if total_pending_individuals > 0:
            fee_summary_text = (
                f"A total of {total_pending_individuals} students attended classes without an active fee subscription "
                f"for {month_str} {year_str}, accumulating {total_pending_visits} attendance days."
            )
        else:
            fee_summary_text = (
                f"✅ Fee Status: Excellent. All {total_attending} attending students currently have active subscriptions "
                f"covering {month_str} {year_str}."
            )
    
    return {
        "presence": presence, "daily_totals": daily_totals, "dow_totals": dow_totals,
        "weekly_totals": weekly_totals, "batch_analysis": batch_analysis,
        "top_attendees": top_attendees, "least_active": least_active,
        "segment_counts": segment_counts, "kpis": kpis, "insights": insights,
        "op_insights": op_insights, "pending_students": pending_students,
        "fee_summary_text": fee_summary_text, "df": df
    }

# ============================================================================
# VISUALIZATION FUNCTIONS
# ============================================================================

def plot_charts(analytics, tmpdir):
    """Generate matplotlib charts for PDF embedding"""
    charts = {}
    
    dow = analytics["dow_totals"]
    if not dow.empty:
        fig, ax = plt.subplots(figsize=(8, 4))
        bars = ax.bar(dow["DayOfWeek"], dow["Attendance"], color=CHART_RED)
        ax.set_title("Attendance by Day of Week", fontweight="bold", fontsize=12)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.set_ylim(0, dow["Attendance"].max() * 1.15)
        
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2, height + dow["Attendance"].max()*0.02,
                   str(int(height)), ha="center", va="bottom", fontweight="bold", fontsize=9)
        
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        dow_path = os.path.join(tmpdir, "dow.png")
        plt.savefig(dow_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        charts["dow"] = dow_path
    
    daily = analytics["daily_totals"]
    if not daily.empty:
        fig, ax = plt.subplots(figsize=(8, 4))
        day_numbers = daily["Date"].dt.strftime("%d")
        ax.plot(day_numbers, daily["Attendance"], marker="o", color=CHART_RED, linewidth=2, markersize=4)
        ax.set_title("Daily Attendance Trend", fontweight="bold", fontsize=12)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.grid(axis='y', linestyle='--', alpha=0.5)
        ax.set_ylim(0, daily["Attendance"].max() * 1.15)
        
        for x, y in zip(day_numbers[::2], daily["Attendance"].values[::2]):
            ax.text(x, y + daily["Attendance"].max()*0.02, str(int(y)), ha="center", fontsize=7)
        
        ax.set_xticks(day_numbers[::2])
        plt.xticks(rotation=0)
        plt.tight_layout()
        trend_path = os.path.join(tmpdir, "trend.png")
        plt.savefig(trend_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        charts["trend"] = trend_path
    
    return charts

# ============================================================================
# PDF GENERATION FUNCTIONS
# ============================================================================

def make_pdf_table(df, headers, widths):
    """Create a styled ReportLab table"""
    if df.empty:
        return Table([["No data available"]], colWidths=widths)
    
    data = [headers] + df.astype(str).values.tolist()
    tbl = Table(data, colWidths=widths, repeatRows=1)
    
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_RED),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_LIGHT]),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    
    return tbl

def build_pdf(out_pdf, meta, analytics, charts):
    """Generate professional PDF report"""
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(name="ReportTitle", fontName="Helvetica-Bold", fontSize=20, leading=24, textColor=BRAND_DARK, spaceAfter=6)
    subtitle_style = ParagraphStyle(name="ReportSubtitle", fontName="Helvetica", fontSize=12, leading=14, textColor=BRAND_GREY, spaceAfter=18)
    section_style = ParagraphStyle(name="SectionHeader", fontName="Helvetica-Bold", fontSize=12, textColor=colors.white, backColor=BRAND_RED, spaceBefore=15, spaceAfter=10, borderPadding=(4, 6, 4, 6))
    bullet_style = ParagraphStyle(name="Bullet", parent=styles["BodyText"], leftIndent=15, spaceAfter=5, bulletIndent=5)
    
    doc = SimpleDocTemplate(out_pdf, pagesize=PAGE_SIZE, rightMargin=RIGHT_MARGIN, leftMargin=LEFT_MARGIN, topMargin=TOP_MARGIN, bottomMargin=BOTTOM_MARGIN, title=f"{meta.center_name} - Attendance Report")
    content = []
    
    content.append(Paragraph(f"{meta.center_name} - Attendance Report", title_style))
    content.append(Paragraph(f"Reporting Period: {meta.report_period}", subtitle_style))
    content.append(Spacer(1, 2 * mm))
    
    content.append(Paragraph("1. Executive Summary", section_style))
    kpi_table_data = [["Metric", "Value"]] + analytics["kpis"].astype(str).values.tolist()
    kpi_t = Table(kpi_table_data, colWidths=[100 * mm, 60 * mm])
    kpi_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_LIGHT]),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"), ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 6)
    ]))
    content.append(kpi_t)
    content.append(Spacer(1, 6 * mm))
    
    content.append(Paragraph("Key Monthly Insights:", ParagraphStyle(name="Sub", fontName="Helvetica-Bold", spaceAfter=6)))
    for ins in analytics["insights"]:
        content.append(Paragraph(f"• {ins}", bullet_style))
    content.append(Spacer(1, 8 * mm))
    
    content.append(Paragraph("2. Attendance Trends & Highlights", section_style))
    for op_ins in analytics["op_insights"]:
        content.append(Paragraph(f"• {op_ins}", bullet_style))
    content.append(Spacer(1, 4 * mm))
    
    week_table_data = [["Weekly Operational Period", "Total Center Visits"]] + analytics["weekly_totals"].astype(str).values.tolist()
    week_t = Table(week_table_data, colWidths=[80 * mm, 60 * mm])
    week_t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_DARK), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_LIGHT]), ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 6)
    ]))
    content.append(week_t)
    content.append(Spacer(1, 6 * mm))
    
    if "dow" in charts:
        content.append(Image(charts["dow"], width=160 * mm, height=70 * mm))
        content.append(Spacer(1, 4 * mm))
    if "trend" in charts:
        content.append(Image(charts["trend"], width=160 * mm, height=70 * mm))
    
    content.append(PageBreak())
    
    content.append(Paragraph("3. Member Insights & Demographics", section_style))
    content.append(Spacer(1, 2 * mm))
    
    top_disp = analytics["top_attendees"].head(10).copy()
    top_disp["StudentName"] = top_disp["StudentName"].apply(lambda x: (str(x)[:14] + '..') if len(str(x)) > 16 else x)
    least_disp = analytics["least_active"].head(10).copy()
    least_disp["StudentName"] = least_disp["StudentName"].apply(lambda x: (str(x)[:14] + '..') if len(str(x)) > 16 else x)
    
    w1 = Table([[
        make_pdf_table(top_disp, ["Top Attendees", "Batch", "Visits"], [45 * mm, 32 * mm, 17 * mm]),
        make_pdf_table(least_disp, ["Least Active", "Batch", "Visits"], [45 * mm, 32 * mm, 17 * mm])
    ]])
    w1.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    
    w2 = Table([[
        make_pdf_table(analytics["segment_counts"], ["Engagement Segment", "Count"], [67 * mm, 27 * mm]),
        make_pdf_table(
            analytics["batch_analysis"].head(5)[["Batch", "Members", "Total_Present"]] if not analytics["batch_analysis"].empty else pd.DataFrame(),
            ["Batch", "Users", "Visits"], [57 * mm, 18 * mm, 19 * mm]
        )
    ]])
    w2.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    
    content.extend([w1, Spacer(1, 8 * mm), w2])
    
    if analytics.get("fee_summary_text"):
        content.append(Spacer(1, 8 * mm))
        content.append(Paragraph("4. Fee Reconciliation: Unpaid Attendees", section_style))
        content.append(Spacer(1, 4 * mm))
        
        summary_style = ParagraphStyle(
            name="FeeSummary", fontName="Helvetica-Bold", fontSize=11,
            textColor=BRAND_DARK if "Excellent" in analytics["fee_summary_text"] else colors.red, spaceAfter=12
        )
        content.append(Paragraph(analytics["fee_summary_text"], summary_style))
        
        if not analytics["pending_students"].empty:
            def_disp = analytics["pending_students"].copy()
            def_disp["StudentName"] = def_disp["StudentName"].apply(lambda x: (str(x)[:16] + '..') if len(str(x)) > 18 else x)
            
            cols = ['StudentName', 'System_ID', 'Days Attended', 'Mobile No', 'Last Payment Date', 'Subscription Coverage']
            headers = ["Student Name", "Student ID", "Days Attended", "Mobile No", "Last Payment", "Coverage"]
            widths = [35 * mm, 25 * mm, 20 * mm, 30 * mm, 25 * mm, 37 * mm]
            
            if 'Status' in def_disp.columns:
                cols.insert(2, 'Status')
                headers.insert(2, "Status")
                widths = [32 * mm, 22 * mm, 18 * mm, 18 * mm, 28 * mm, 22 * mm, 32 * mm]
            
            content.append(make_pdf_table(def_disp[cols], headers, widths))
    
    def draw_page_setup(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(BRAND_RED)
        canvas.setLineWidth(1)
        canvas.line(LEFT_MARGIN, BOTTOM_MARGIN - 5 * mm, A4[0] - RIGHT_MARGIN, BOTTOM_MARGIN - 5 * mm)
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(BRAND_GREY)
        canvas.drawString(LEFT_MARGIN, BOTTOM_MARGIN - 10 * mm, "Quality and Systems, Rashtrotthana Parishat")
        canvas.drawRightString(A4[0] - RIGHT_MARGIN, BOTTOM_MARGIN - 10 * mm, f"Page {doc.page}")
        canvas.restoreState()
    
    doc.build(content, onFirstPage=draw_page_setup, onLaterPages=draw_page_setup)

# ============================================================================
# EXCEL GENERATION FUNCTIONS
# ============================================================================

def write_excel_report(out_xlsx, meta, df, analytics, charts):
    """Generate comprehensive Excel audit file"""
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        analytics["kpis"].to_excel(writer, sheet_name="Summary", index=False, startrow=4)
        
        all_insights = analytics["insights"] + analytics["op_insights"]
        if analytics.get("fee_summary_text"):
            all_insights.append(analytics["fee_summary_text"])
        
        pd.DataFrame({"Insight": all_insights}).to_excel(
            writer, sheet_name="Summary", index=False, startrow=4 + len(analytics["kpis"]) + 3
        )
        
        if not analytics["pending_students"].empty:
            analytics["pending_students"].to_excel(writer, sheet_name="Pending_Renewals_List", index=False)
        
        analytics["segment_counts"].to_excel(writer, sheet_name="Member_Segments", index=False)
        analytics["batch_analysis"].to_excel(writer, sheet_name="Batch_Analysis", index=False)
        analytics["weekly_totals"].to_excel(writer, sheet_name="Weekly_Trends", index=False)
        analytics["daily_totals"].to_excel(writer, sheet_name="Daily_Attendance", index=False)
        analytics["dow_totals"].to_excel(writer, sheet_name="Day_of_Week", index=False)
        analytics["top_attendees"].head(50).to_excel(writer, sheet_name="Top_Attendees", index=False)
        analytics["least_active"].head(50).to_excel(writer, sheet_name="Least_Active", index=False)
        df.drop(columns=['System_ID'], errors='ignore').to_excel(writer, sheet_name="Clean_Data", index=False)
    
    wb = load_workbook(out_xlsx)
    for ws in wb.worksheets:
        for c_cell in ws[1]:
            c_cell.font = Font(bold=True, color="FFFFFF")
            c_cell.fill = PatternFill("solid", fgColor="8A1E1E")
            c_cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
    
    ws = wb["Summary"]
    ws["A1"], ws["B1"], ws["A2"], ws["B2"], ws["A3"], ws["B3"] = \
        "Center", meta.center_name, "Month", meta.report_month, "Period", meta.report_period
    
    for cell in ["A1", "A2", "A3"]:
        ws[cell].font = Font(bold=True)
    
    try:
        if "dow" in charts:
            ws.add_image(XLImage(charts["dow"]), "E2")
        if "trend" in charts:
            ws.add_image(XLImage(charts["trend"]), "E22")
    except Exception:
        pass
    
    wb.save(out_xlsx)

# ============================================================================
# STREAMLIT APPLICATION
# ============================================================================

def main():
    st.set_page_config(page_title="Yoga Center Operations Manager", page_icon="🧘", layout="centered")
    
    st.title("🧘 Yoga Center Operations Manager")
    st.markdown("""
    **Upload your monthly attendance to generate the standard report.**  
    **Upload the Fee Report alongside it to activate the Fee Reconciliation checks.**
    """)
    
    with st.expander("📋 Expected File Format (Click to View)"):
        st.markdown("""
        **Attendance File Must Have:**
        - Columns: `StudentId` (e.g., `9-GYN/Y/4162`), `StudentName`, `Batch`, `Center`, `Status`
        - Date columns: `5/1/26` to `5/31/26` (or similar date format)
        - Values in date cells: `Present`, `Absent`, or blank
        
        **Fee File Must Have:**
        - Columns: `Stud ID` (numeric), `Stud Name`, `Mobile No`, `Particulars`, `Status`, `Pymt Dt`
        - `Particulars` should contain month coverage like `[May-26]`, `[Jun-26]`
        - `Status`: `Active` or `Canceled`
        """)
    
    col1, col2 = st.columns(2)
    with col1:
        uploaded_att = st.file_uploader("1. Required: Upload Attendance Report (.xlsx, .csv)", type=["xlsx", "csv"], key="attendance")
    with col2:
        uploaded_fee = st.file_uploader("2. Optional: Upload Fee Report (.xlsx, .csv)", type=["xlsx", "csv"], key="fee")
    
    if uploaded_att is not None:
        if st.button("Generate Reports ⚙️", type="primary", use_container_width=True):
            with st.status("Processing files...", expanded=True) as status:
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        st.write("✓ Loading attendance data...")
                        if uploaded_att.name.endswith('.csv'):
                            att_path = os.path.join(tmpdir, "input.csv")
                            with open(att_path, "wb") as f: f.write(uploaded_att.getvalue())
                            pd.read_csv(att_path, header=None).to_excel(os.path.join(tmpdir, "input.xlsx"), index=False, header=False)
                            att_path = os.path.join(tmpdir, "input.xlsx")
                        else:
                            att_path = os.path.join(tmpdir, "input.xlsx")
                            with open(att_path, "wb") as f: f.write(uploaded_att.getvalue())
                        
                        fee_df = None
                        if uploaded_fee:
                            st.write("✓ Parsing fee records...")
                            if uploaded_fee.name.endswith('.csv'):
                                fee_path = os.path.join(tmpdir, "fee.csv")
                                with open(fee_path, "wb") as f: f.write(uploaded_fee.getvalue())
                                pd.read_csv(fee_path, header=None).to_excel(os.path.join(tmpdir, "fee.xlsx"), index=False, header=False)
                                fee_path = os.path.join(tmpdir, "fee.xlsx")
                            else:
                                fee_path = os.path.join(tmpdir, "fee.xlsx")
                                with open(fee_path, "wb") as f: f.write(uploaded_fee.getvalue())
                            
                            fee_df = read_fee_report(fee_path)
                            st.write(f"  → Loaded {len(fee_df)} fee records")
                        
                        st.write("✓ Analyzing attendance data...")
                        df, date_cols, date_idx, meta, presence = read_and_clean_attendance(att_path)
                        st.write(f"  → Found {len(df)} students across {len(date_cols)} days")
                        
                        st.write("✓ Reconciling attendance vs fees...")
                        analytics = build_analytics(df, date_cols, date_idx, presence, fee_df, meta)
                        
                        st.write("✓ Creating visualizations...")
                        charts = plot_charts(analytics, tmpdir)
                        
                        st.write("✓ Building PDF dashboard...")
                        pdf_path = os.path.join(tmpdir, "Report.pdf")
                        build_pdf(pdf_path, meta, analytics, charts)
                        
                        st.write("✓ Building Excel audit...")
                        xlsx_path = os.path.join(tmpdir, "Audit.xlsx")
                        write_excel_report(xlsx_path, meta, df, analytics, charts)
                        
                        with open(pdf_path, "rb") as f: pdf_bytes = f.read()
                        with open(xlsx_path, "rb") as f: xlsx_bytes = f.read()
                    
                    status.update(label="✅ Reports generated successfully!", state="complete")
                    
                    if fee_df is None:
                        success_msg = f"✅ Standard attendance report generated for **{meta.center_name}**."
                    else:
                        unpaid_count = len(analytics["pending_students"])
                        if unpaid_count > 0:
                            success_msg = f"⚠️ Report generated for **{meta.center_name}**. Found **{unpaid_count} unpaid attendees**."
                        else:
                            success_msg = f"✅ Report generated for **{meta.center_name}**. All attendees have valid fees!"
                    
                    st.success(success_msg)
                    
                    col1, col2 = st.columns(2)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                    pdf_name = f"{meta.center_name.replace(' ', '_')}_Report_{meta.month_key}_{timestamp}.pdf"
                    xlsx_name = f"{meta.center_name.replace(' ', '_')}_Audit_{meta.month_key}_{timestamp}.xlsx"
                    
                    with col1:
                        st.download_button(label="📄 Download PDF Dashboard", data=pdf_bytes, file_name=pdf_name, mime="application/pdf", use_container_width=True)
                    with col2:
                        st.download_button(label="📊 Download Excel Audit", data=xlsx_bytes, file_name=xlsx_name, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    
                    if st.checkbox("🔍 Preview Data (Optional)", value=False):
                        st.subheader("Attendance Sample")
                        st.dataframe(df[['StudentName', 'Batch', 'Present_Count_Calc']].head(10))
                        if fee_df is not None:
                            st.subheader("Fee Records Sample")
                            st.dataframe(fee_df[['Norm_ID', 'Particulars', 'Status']].head(10))
                
                except ReportValidationError as ve:
                    status.update(label="❌ Validation Error", state="error")
                    st.warning(str(ve))
                except Exception as e:
                    status.update(label="❌ Error Occurred", state="error")
                    st.error("❌ An error occurred during report generation.")
                    st.exception(e)
    else:
        st.info("👆 Please upload an attendance file to begin.")
    
    st.markdown("---")
    st.markdown("""
    **Need Help?**  
    - Ensure Student IDs match between attendance and fee files
    - Fee file must have month coverage in `Particulars` column (e.g., `[May-26]`)
    - Contact support if you encounter persistent errors
    """)

if __name__ == "__main__":
    main()
